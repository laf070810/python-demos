import itchat
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import re
import sys
from PyQt5.QtWidgets import QApplication, QMainWindow
from window import Ui_MainWindow


class Window(Ui_MainWindow, QMainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.show()

        self.button.clicked.connect(self.login)

    def login(self):
        itchat.auto_login(enableCmdQR=False, hotReload=False)
        self.chatroom_list.clear()
        for chatroom in itchat.get_chatrooms():
            self.chatroom_list.addItem(chatroom['NickName'])
        self.chatroom_list.setEnabled(True)
        self.button.setText('导出')
        self.button.clicked.disconnect(self.login)
        self.button.clicked.connect(self.export)
        itchat.run(blockThread=False)

    def export(self):
        wb = Workbook()
        wb.remove_sheet(wb.active)

        for chatroom in itchat.search_chatrooms(name=self.chatroom_list.currentText()):
            ws = wb.create_sheet(chatroom['NickName'].replace('*', ''))
            ws['A1'] = '用户昵称'
            ws['B1'] = '备注名'
            ws['C1'] = '群昵称'
            ws['D1'] = '学号'
            ws['A1'].font = Font(bold=True)
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B1'].font = Font(bold=True)
            ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C1'].font = Font(bold=True)
            ws['C1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D1'].font = Font(bold=True)
            ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            current_row = 2

            for member in itchat.update_chatroom(chatroom['UserName'], detailedMember=True)['MemberList']:
                ws.cell(row=current_row, column=1, value=member['NickName'])
                ws.cell(row=current_row, column=2, value=member['RemarkName'])
                ws.cell(row=current_row, column=3, value=member['DisplayName'])
                result = re.search('\d{10}', member['DisplayName'])
                if result is not None:
                    ws.cell(row=current_row, column=4, value=int(result.group()))
                current_row += 1

        wb.save('output.xlsx')


app = QApplication(sys.argv)
a = Window()
sys.exit(app.exec_())


'''sample retval from itchat.get_chatrooms()
[<Chatroom: {'MemberList': <ContactList: []>, 'Uin': 0, 'UserName': '@@082681f1ae4075b88be3e144bc053e5a64f3b251a121a7550b21c082eec2cb44', 'NickName': '历届封开大学生协会交流群', 'HeadImgUrl': '/cgi-bin/mmwebwx-bin/webwxgetheadimg?seq=680790900&username=@@082681f1ae4075b88be3e144bc053e5a64f3b251a121a7550b21c082eec2cb44&skey=@crypt_70e610f4_19dc7a1d736918a6b674a511b77ddd27', 'ContactFlag': 2, 'MemberCount': 199, 'RemarkName': '', 'HideInputBarFlag': 0, 'Sex': 0, 'Signature': '', 'VerifyFlag': 0, 'OwnerUin': 0, 'PYInitial': '', 'PYQuanPin': '', 'RemarkPYInitial': '', 'RemarkPYQuanPin': '', 'StarFriend': 0, 'AppAccountFlag': 0, 'Statues': 0, 'AttrStatus': 0, 'Province': '', 'City': '', 'Alias': '', 'SnsFlag': 0, 'UniFriend': 0, 'DisplayName': '', 'ChatRoomId': 0, 'KeyWord': '', 'EncryChatRoomId': '', 'IsOwner': 0, 'IsAdmin': None, 'Self': <User: {'MemberList': <ContactList: []>, 'UserName': '@a920724d87b302a01cecbed712b20a3b16a797b0fc95a6e3df0a809b86daa1bb', 'City': '', 'DisplayName': '', 'PYQuanPin': '', 'RemarkPYInitial': '', 'Province': '', 'KeyWord': '', 'RemarkName': '', 'PYInitial': '', 'EncryChatRoomId': '', 'Alias': '', 'Signature': '', 'NickName': '木子。', 'RemarkPYQuanPin': '', 'HeadImgUrl': '/cgi-bin/mmwebwx-bin/webwxgeticon?seq=1878403289&username=@a920724d87b302a01cecbed712b20a3b16a797b0fc95a6e3df0a809b86daa1bb&skey=@crypt_70e610f4_19dc7a1d736918a6b674a511b77ddd27', 'UniFriend': 0, 'Sex': 1, 'AppAccountFlag': 0, 'VerifyFlag': 0, 'ChatRoomId': 0, 'HideInputBarFlag': 0, 'AttrStatus': 0, 'SnsFlag': 17, 'MemberCount': 0, 'OwnerUin': 0, 'ContactFlag': 0, 'Uin': 2987934061, 'StarFriend': 0, 'Statues': 0, 'WebWxPluginSwitch': 0, 'HeadImgFlag': 1}>}>, 
<Chatroom: {'MemberList': <ContactList: []>, 'Uin': 0, 'UserName': '@@fc8da34f74d221b9a39c7d96d5f58d1fc4a971b5ca89ac99c095e45d71e8365e', 'NickName': '2019复分析学习群', 'HeadImgUrl': '/cgi-bin/mmwebwx-bin/webwxgetheadimg?seq=0&username=@@fc8da34f74d221b9a39c7d96d5f58d1fc4a971b5ca89ac99c095e45d71e8365e&skey=@crypt_70e610f4_19dc7a1d736918a6b674a511b77ddd27', 'ContactFlag': 0, 'MemberCount': 105, 'RemarkName': '', 'HideInputBarFlag': 0, 'Sex': 0, 'Signature': '', 'VerifyFlag': 0, 'OwnerUin': 0, 'PYInitial': '', 'PYQuanPin': '', 'RemarkPYInitial': '', 'RemarkPYQuanPin': '', 'StarFriend': 0, 'AppAccountFlag': 0, 'Statues': 1, 'AttrStatus': 0, 'Province': '', 'City': '', 'Alias': '', 'SnsFlag': 0, 'UniFriend': 0, 'DisplayName': '', 'ChatRoomId': 0, 'KeyWord': '', 'EncryChatRoomId': '', 'IsOwner': 0, 'IsAdmin': None, 'Self': <User: {'MemberList': <ContactList: []>, 'UserName': '@a920724d87b302a01cecbed712b20a3b16a797b0fc95a6e3df0a809b86daa1bb', 'City': '', 'DisplayName': '', 'PYQuanPin': '', 'RemarkPYInitial': '', 'Province': '', 'KeyWord': '', 'RemarkName': '', 'PYInitial': '', 'EncryChatRoomId': '', 'Alias': '', 'Signature': '', 'NickName': '木子。', 'RemarkPYQuanPin': '', 'HeadImgUrl': '/cgi-bin/mmwebwx-bin/webwxgeticon?seq=1878403289&username=@a920724d87b302a01cecbed712b20a3b16a797b0fc95a6e3df0a809b86daa1bb&skey=@crypt_70e610f4_19dc7a1d736918a6b674a511b77ddd27', 'UniFriend': 0, 'Sex': 1, 'AppAccountFlag': 0, 'VerifyFlag': 0, 'ChatRoomId': 0, 'HideInputBarFlag': 0, 'AttrStatus': 0, 'SnsFlag': 17, 'MemberCount': 0, 'OwnerUin': 0, 'ContactFlag': 0, 'Uin': 2987934061, 'StarFriend': 0, 'Statues': 0, 'WebWxPluginSwitch': 0, 'HeadImgFlag': 1}>}>,
'''

'''sample retval from itchat.update_chatroom()
{'MemberList': <ContactList:
[<ChatroomMember: {'MemberList': <ContactList: []>, 'Uin': 0, 'UserName': '@bdbe4330bb8b903746942bae9f6ff960', 'NickName': 'Sea🍂', 'HeadImgUrl': '/cgi-bin/mmwebwx-bin/webwxgeticon?seq=0&username=@bdbe4330bb8b903746942bae9f6ff960&chatroomid=@f24e5a8c2a41a24140920f7ecc71dd68&skey=', 'ContactFlag': 0, 'MemberCount': 0, 'RemarkName': '', 'HideInputBarFlag': 0, 'Sex': 2, 'Signature': '对热爱的东西，别敷衍。', 'VerifyFlag': 0, 'OwnerUin': 0, 'PYInitial': 'SEASPANCLASSEMOJIEMOJI1F342SPAN', 'PYQuanPin': 'Seaspanclassemojiemoji1f342span', 'RemarkPYInitial': '', 'RemarkPYQuanPin': '', 'StarFriend': 0, 'AppAccountFlag': 0, 'Statues': 0, 'AttrStatus': 102503, 'Province': '广东', 'City': '广州', 'Alias': '', 'SnsFlag': 17, 'UniFriend': 0, 'DisplayName': '', 'ChatRoomId': 0, 'KeyWord': 'lin', 'EncryChatRoomId': '@f24e5a8c2a41a24140920f7ecc71dd68', 'IsOwner': 0}>, 
'''
